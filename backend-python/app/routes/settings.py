import shutil,sqlite3,secrets,csv,io
from datetime import datetime,timezone
from pathlib import Path
from zoneinfo import ZoneInfo,ZoneInfoNotFoundError
from fastapi import APIRouter,Depends,File,HTTPException,UploadFile
from fastapi.responses import FileResponse,StreamingResponse
from ..audit import log_audit
from ..database import active_db_path,fetch_all,fetch_one,transaction
from ..security import User,roles
from ..stock import receive
from ..backup_service import backup_schedule,verify_backup,scheduler_tick
router=APIRouter(prefix='/api/settings',tags=['settings']);admin=roles('SupplyChainManager');ROOT=Path(__file__).resolve().parents[3];UPLOADS=Path(__file__).resolve().parents[2]/'uploads';BACKUPS=ROOT/'backend-python'/'backups';BACKUPS.mkdir(exist_ok=True)
@router.get('')
def all_settings(_u:User):return {x['key']:x['value']for x in fetch_all('SELECT key,value FROM settings')}
@router.get('/approval-limits')
def limits(_u:dict=Depends(roles('SupplyChainManager','PurchaseManager','WarehouseManager'))):return {x['key']:float(x['value'])for x in fetch_all("SELECT key,value FROM settings WHERE key LIKE 'approval_limit_%' OR key='material_issue_approval_threshold'")}
@router.get('/company')
def company(_u:User):return fetch_one('SELECT * FROM company WHERE deleted_at IS NULL ORDER BY id DESC LIMIT 1')or{}
@router.get('/branding')
def branding():
    row=fetch_one('SELECT name,logo_url,address,phone,email,website,tax_info,registration_number,branch_info,currency,base_currency,country_code,time_zone,financial_year FROM company WHERE deleted_at IS NULL ORDER BY id DESC LIMIT 1')or{}
    return {'company_name':row.get('name')or'Company Name','logo_url':row.get('logo_url'),'address':row.get('address')or'','phone':row.get('phone')or'','email':row.get('email')or'','website':row.get('website')or'','tax_info':row.get('tax_info')or'','registration_number':row.get('registration_number')or'','branch_info':row.get('branch_info')or'','currency':row.get('base_currency')or row.get('currency')or'SAR','base_currency':row.get('base_currency')or row.get('currency')or'SAR','country_code':row.get('country_code')or'SA','time_zone':row.get('time_zone')or'Asia/Riyadh','financial_year':row.get('financial_year')or'','application_name':'ProcuraFlow'}
@router.put('/company')
def update_company(body:dict,user:dict=Depends(admin)):
    allowed=['name','address','phone','email','website','registration_number','branch_info','tax_info','currency','base_currency','country_code','city_id','postal_code','region_province','financial_year','time_zone'];keys=[k for k in body if k in allowed]
    if not keys:raise HTTPException(400,'No company fields provided')
    row=fetch_one('SELECT * FROM company WHERE deleted_at IS NULL ORDER BY id DESC LIMIT 1')
    with transaction(immediate=True)as c:c.execute(f"UPDATE company SET {','.join(k+'=?'for k in keys)} WHERE id=?",tuple(body[k]for k in keys)+(row['id'],));log_audit(c,'company',row['id'],'UPDATE',user['id'],row,body)
    return fetch_one('SELECT * FROM company WHERE id=?',(row['id'],))
@router.put('/{key}')
def set_value(key:str,body:dict,user:User):
    management=['SupplyChainManager'];
    if key=='global_transport_modes':management+=['WarehouseManager','WarehouseSupervisor']
    if user['role']not in management:raise HTTPException(403,'Only an authorized manager may change this setting')
    value=body.get('value')
    with transaction(immediate=True)as c:c.execute('INSERT INTO settings(key,value)VALUES(?,?)ON CONFLICT(key)DO UPDATE SET value=excluded.value',(key,str(value)));log_audit(c,'settings',None,'UPDATE',user['id'],after={'key':key,'value':value})
    return {'key':key,'value':value}
@router.post('/backup')
def backup(user:dict=Depends(admin)):
    name=f"procuraflow-{secrets.token_hex(6)}.db";target=BACKUPS/name
    source=sqlite3.connect(active_db_path());counts={table:source.execute(f'SELECT COUNT(*) FROM {table}').fetchone()[0]for table in('users','employees','items','warehouses','inventory_stock','inventory_layers','stock_ledger','audit_log')};destination=sqlite3.connect(target);source.backup(destination);destination.close();source.close();verification=verify_backup(target,counts)
    with transaction(immediate=True)as c:c.execute("INSERT INTO backup_restore_history(backup_reference,backup_type,database_included,attachments_included,configuration_included,backup_status,restore_tested,restore_test_date,restore_result,notes)VALUES(?,'MANUAL',1,0,1,'SUCCESS',1,datetime('now'),?,'Verified FastAPI SQLite backup')",(name,verification));log_audit(c,'backup_restore_history',None,'CREATE',user['id'],after={'backup_reference':name,'verification':verification})
    return FileResponse(target,media_type='application/octet-stream',filename=name)

@router.get('/maintenance/status')
def maintenance_status(warehouse_id:int|None=None):
    state=fetch_one('SELECT active_yn,reason,cycle_id,scheduled_at_utc,started_at,completed_at,result FROM system_maintenance WHERE id=1')or{}
    schedule=backup_schedule();scheduled=schedule['scheduled_utc'];local_zone=schedule['time_zone']
    if warehouse_id:
        warehouse=fetch_one('SELECT time_zone FROM warehouses WHERE id=? AND deleted_at IS NULL',(warehouse_id,))
        if warehouse:
            try:local_zone=warehouse['time_zone'];scheduled=scheduled.astimezone(ZoneInfo(local_zone))
            except ZoneInfoNotFoundError:pass
    last=fetch_one("SELECT backup_status,created_at,restore_result FROM backup_restore_history WHERE backup_status='SUCCESS' ORDER BY id DESC LIMIT 1")or{}
    return {**state,'active_yn':bool(state.get('active_yn')),'automatic_enabled':schedule['enabled'],'scheduled_at':scheduled.isoformat(),'system_scheduled_at_utc':schedule['scheduled_utc'].isoformat(),'display_time_zone':local_zone,'seconds_remaining':max(0,int((schedule['scheduled_utc']-datetime.now(timezone.utc)).total_seconds())),'warning_active':0<(schedule['scheduled_utc']-datetime.now(timezone.utc)).total_seconds()<=1800,'last_successful_backup':last}

@router.get('/backup-policy')
def backup_policy(_user:dict=Depends(admin)):
    keys=('automatic_month_end_backup','backup_time','backup_time_zone','backup_warning_minutes','backup_reminder_minutes')
    return {key:(fetch_one('SELECT value FROM settings WHERE key=?',(key,))or{}).get('value')for key in keys}

@router.put('/backup-policy')
def update_backup_policy(body:dict,user:dict=Depends(admin)):
    allowed={'automatic_month_end_backup','backup_time','backup_time_zone','backup_warning_minutes','backup_reminder_minutes'}
    unknown=set(body)-allowed
    if unknown:raise HTTPException(400,f"Unknown backup settings: {', '.join(sorted(unknown))}")
    try:ZoneInfo(str(body.get('backup_time_zone')or backup_schedule()['time_zone']))
    except ZoneInfoNotFoundError:raise HTTPException(400,'Select a valid IANA backup time zone')
    if body.get('backup_time'):
        try:h,m=map(int,str(body['backup_time']).split(':'));assert 0<=h<24 and 0<=m<60
        except(ValueError,AssertionError):raise HTTPException(400,'Backup time must use HH:MM format')
    with transaction(immediate=True)as c:
        for key,value in body.items():c.execute('INSERT INTO settings(key,value)VALUES(?,?)ON CONFLICT(key)DO UPDATE SET value=excluded.value',(key,str(value)))
        log_audit(c,'settings',None,'UPDATE',user['id'],after={'backup_policy':body})
    return backup_policy(user)
@router.post('/restore')
async def restore(user:dict=Depends(admin),backup:UploadFile=File(...)):
    data=await backup.read();stage=BACKUPS/f"restore-stage-{secrets.token_hex(6)}.db";stage.write_bytes(data)
    try:
        c=sqlite3.connect(stage);result=c.execute('PRAGMA integrity_check').fetchone()[0];c.close()
        if result!='ok':raise HTTPException(400,'Backup database failed integrity validation')
    except sqlite3.Error:stage.unlink(missing_ok=True);raise HTTPException(400,'Uploaded file is not a valid SQLite backup')
    return {'success':True,'status':'Validated and staged','staged_file':stage.name,'message':'Restart with DB_PATH pointing to the staged database after controlled approval.'}
@router.post('/factory-reset')
def factory_reset(_user:dict=Depends(admin)):raise HTTPException(409,'Factory reset requires the dedicated offline maintenance command to prevent accidental live data destruction')
@router.post('/fiscal-close')
def fiscal_close(body:dict,user:dict=Depends(admin)):
    with transaction(immediate=True)as c:c.execute("UPDATE accounting_periods SET status='CLOSED',closed_by=?,closed_at=datetime('now'),reason=? WHERE fiscal_year=? AND status<>'CLOSED'",(user['id'],body.get('reason')or'Fiscal close',body.get('fiscal_year')));changed=c.total_changes
    return {'success':True,'closed_periods':changed}
@router.post('/company-logo')
async def company_logo(user:dict=Depends(admin),logo:UploadFile=File(...)):
    data=await logo.read(5*1024*1024+1)
    valid_png=logo.content_type=='image/png'and data[:8]==b'\x89PNG\r\n\x1a\n';valid_jpeg=logo.content_type=='image/jpeg'and data[:3]==b'\xff\xd8\xff'
    if len(data)>5*1024*1024 or not(valid_png or valid_jpeg):raise HTTPException(400,'Select a valid PNG or JPG logo no larger than 5 MB')
    active=fetch_one('SELECT id,logo_url FROM company WHERE deleted_at IS NULL ORDER BY id DESC LIMIT 1')
    if not active:raise HTTPException(404,'Active company record not found')
    ext='.png'if logo.content_type=='image/png'else'.jpg';folder=UPLOADS/'logos';folder.mkdir(parents=True,exist_ok=True);name='company-logo-'+secrets.token_hex(6)+ext;(folder/name).write_bytes(data);url='/uploads/logos/'+name
    try:
        with transaction(immediate=True)as c:c.execute('UPDATE company SET logo_url=? WHERE id=?',(url,active['id']));log_audit(c,'company',active['id'],'UPDATE',user['id'],{'logo_url':active.get('logo_url')},{'logo_url':url})
    except Exception:
        (folder/name).unlink(missing_ok=True);raise
    previous=str(active.get('logo_url')or'')
    if previous.startswith('/uploads/logos/'):
        old=folder/Path(previous).name
        if old.name!=name:old.unlink(missing_ok=True)
    return {'logo_url':url}
def rows_from_upload(data,name):
    if name.lower().endswith('.csv'):return list(csv.DictReader(io.StringIO(data.decode('utf-8-sig'))))
    from openpyxl import load_workbook
    book=load_workbook(io.BytesIO(data),read_only=True,data_only=True);sheet=book.active;values=list(sheet.values)
    header_index=next((index for index,row in enumerate(values)if sum(value not in(None,'')for value in row)>=2),None)
    if header_index is None:return []
    headers=[str(x or'').strip()for x in values[header_index]];return[dict(zip(headers,row))for row in values[header_index+1:]if any(x is not None for x in row)]
ITEM_IMPORT_FIELDS=['description','category','subcategory','uom','purchase_uom','issue_uom','conversion_factor','consumable_returnable','high_value_flag','always_approval_yn','tool_control_yn','batch_control_yn','expiry_control_yn','inspection_required_yn','min_stock','max_stock','reorder_level','standard_cost']
IMPORT_TEMPLATES={
    'vendors':(['name','contact_person','phone','email','address','payment_terms'],['Example Supplier','John Doe','+971501234567','contact@example.com','Warehouse City','Net 30']),
    'items':(ITEM_IMPORT_FIELDS,['Example Steel Bar','Raw Material','Steel','KG','TON','KG',1000,'Consumable',0,0,0,1,0,1,100,1000,250,12.5]),
    'opening-balances':(ITEM_IMPORT_FIELDS+['warehouse','location','quantity','unit_cost','received_date','batch','expiry_date'],['Example Steel Bar','Raw Material','Steel','KG','TON','KG',1000,'Consumable',0,0,0,1,0,1,100,1000,250,12.5,'WH-001','WH-001-BN-0003',100,5.2,'2026-01-01','BATCH-001','']),
}
@router.get('/imports/{template_type}/template')
def download_import_template(template_type:str,user:dict=Depends(admin)):
    template=IMPORT_TEMPLATES.get(template_type)
    if not template:raise HTTPException(404,'Unknown import template')
    from openpyxl import Workbook
    from openpyxl.styles import Alignment,Font,PatternFill,Border,Side
    book=Workbook();sheet=book.active;sheet.title='Import Template';sheet.append(['ProcuraFlow - Precast Supply Chain Control System']);sheet.append([f'{template_type.replace("-"," ").title()} Import Template']);sheet.append(template[0]);sheet.append(template[1]);sheet.freeze_panes='A4';sheet.auto_filter.ref=f'A3:{sheet.cell(3,len(template[0])).column_letter}4'
    sheet.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(template[0]));sheet.merge_cells(start_row=2,start_column=1,end_row=2,end_column=len(template[0]))
    sheet['A1'].font=Font(bold=True,color='FFFFFF',size=16);sheet['A1'].fill=PatternFill('solid',fgColor='052F5F');sheet['A1'].alignment=Alignment(vertical='center')
    sheet['A2'].font=Font(bold=True,color='075FA8',size=12);sheet['A2'].fill=PatternFill('solid',fgColor='F2F7FA')
    for cell in sheet[3]:cell.font=Font(bold=True,color='FFFFFF');cell.fill=PatternFill('solid',fgColor='075FA8');cell.alignment=Alignment(wrap_text=True,vertical='center');cell.border=Border(bottom=Side(style='thin',color='0796A5'))
    sheet.row_dimensions[1].height=28;sheet.row_dimensions[2].height=22;sheet.row_dimensions[3].height=30
    for index,header in enumerate(template[0],1):sheet.column_dimensions[sheet.cell(3,index).column_letter].width=min(34,max(12,len(header)+3))
    output=io.BytesIO();book.save(output);output.seek(0);filename=f'{template_type}-template.xlsx'
    return StreamingResponse(output,media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',headers={'Content-Disposition':f'attachment; filename="{filename}"'})
def normalized_row(row):return {str(key or'').strip().lower():value for key,value in row.items()}
def item_import_values(row,row_number):
    row=normalized_row(row);missing=[field for field in ITEM_IMPORT_FIELDS if row.get(field)in(None,'')]
    if missing:raise HTTPException(400,f"Row {row_number}: mandatory fields missing: {', '.join(missing)}")
    text={key:str(row[key]).strip()for key in('description','category','subcategory','uom','purchase_uom','issue_uom','consumable_returnable')}
    if text['consumable_returnable']not in('Consumable','Returnable'):raise HTTPException(400,f'Row {row_number}: consumable_returnable must be Consumable or Returnable')
    values={**text}
    try:
        for key in('conversion_factor','min_stock','max_stock','reorder_level','standard_cost'):values[key]=float(row[key])
        for key in('high_value_flag','always_approval_yn','tool_control_yn','batch_control_yn','expiry_control_yn','inspection_required_yn'):
            raw=str(row[key]).strip().lower()
            if raw not in('0','1','yes','no','true','false'):raise ValueError
            values[key]=int(raw in('1','yes','true'))
    except(ValueError,TypeError):raise HTTPException(400,f'Row {row_number}: numeric fields and control flags contain invalid data')
    if values['conversion_factor']<=0 or min(values['min_stock'],values['max_stock'],values['reorder_level'],values['standard_cost'])<0:raise HTTPException(400,f'Row {row_number}: conversion must be positive and stock/cost values cannot be negative')
    if values['max_stock']and values['min_stock']>values['max_stock']:raise HTTPException(400,f'Row {row_number}: minimum stock cannot exceed maximum stock')
    return row,values
def next_item_code(connection):
    rows=connection.execute("SELECT item_code code FROM items WHERE item_code LIKE 'ITM-%'").fetchall();numbers=[int(str(row['code']).rsplit('-',1)[-1])for row in rows if str(row['code']).rsplit('-',1)[-1].isdigit()]
    return f"ITM-{(max(numbers)if numbers else 0)+1:04d}"
def ensure_taxonomy(connection,category,subcategory,user_id):
    connection.execute('INSERT OR IGNORE INTO item_categories(name,created_by)VALUES(?,?)',(category,user_id));category_id=connection.execute('SELECT id FROM item_categories WHERE name=? COLLATE NOCASE',(category,)).fetchone()['id'];connection.execute('INSERT OR IGNORE INTO item_subcategories(category_id,name,created_by)VALUES(?,?,?)',(category_id,subcategory,user_id))
def insert_import_item(connection,values,user_id):
    code=next_item_code(connection);keys=list(values);cursor=connection.execute(f"INSERT INTO items(item_code,{','.join(keys)})VALUES(?,{','.join('?'for _ in keys)})",(code,*[values[key]for key in keys]));ensure_taxonomy(connection,values['category'],values['subcategory'],user_id);log_audit(connection,'items',cursor.lastrowid,'CREATE',user_id,after={**values,'item_code':code,'source':'IMPORT'});return cursor.lastrowid
@router.post('/imports/vendors')
async def import_vendors(user:dict=Depends(admin),file:UploadFile=File(...)):
    rows=rows_from_upload(await file.read(),file.filename or'')
    if not rows:raise HTTPException(400,'The supplier import file contains no data rows')
    required=('name','contact_person','phone','email','address','payment_terms');validated=[]
    for index,source in enumerate(rows,2):
        row=normalized_row(source);missing=[field for field in required if row.get(field)in(None,'')]
        if missing:raise HTTPException(400,f"Row {index}: mandatory supplier fields missing: {', '.join(missing)}")
        validated.append({field:str(row[field]).strip()for field in required})
    with transaction(immediate=True)as c:
        existing_codes=[int(str(row['supplier_code']).rsplit('-',1)[-1])for row in c.execute("SELECT supplier_code FROM suppliers WHERE supplier_code LIKE 'SUP-%'").fetchall()if str(row['supplier_code']).rsplit('-',1)[-1].isdigit()];sequence=max(existing_codes)if existing_codes else 0
        for index,row in enumerate(validated,2):
            if c.execute('SELECT id FROM suppliers WHERE deleted_at IS NULL AND lower(trim(name))=lower(trim(?))',(row['name'],)).fetchone():raise HTTPException(409,f"Row {index}: supplier already exists")
            sequence+=1;code=f'SUP-{sequence:04d}';supplier_id=c.execute('INSERT INTO suppliers(supplier_code,name,contact_person,phone,email,address,payment_terms,rating)VALUES(?,?,?,?,?,?,?,0)',(code,row['name'],row['contact_person'],row['phone'],row['email'],row['address'],row['payment_terms'])).lastrowid;log_audit(c,'suppliers',supplier_id,'CREATE',user['id'],after={**row,'supplier_code':code,'source':'IMPORT'})
    return {'imported':len(validated),'created':len(validated),'rows':len(rows)}
@router.post('/imports/items')
async def import_items(user:dict=Depends(admin),file:UploadFile=File(...)):
    rows=rows_from_upload(await file.read(),file.filename or'')
    if not rows:raise HTTPException(400,'The item import file contains no data rows')
    validated=[item_import_values(row,index)for index,row in enumerate(rows,2)]
    with transaction(immediate=True)as c:
        for index,(_row,values) in enumerate(validated,2):
            if c.execute('SELECT id FROM items WHERE deleted_at IS NULL AND lower(trim(description))=lower(trim(?))',(values['description'],)).fetchone():raise HTTPException(409,f"Row {index}: item description already exists; use Item Master to review the existing record")
            insert_import_item(c,values,user['id'])
        log_audit(c,'items',0,'CREATE',user['id'],after={'source':'ITEM_IMPORT','rows':len(validated)})
    return {'imported':len(validated),'created':len(validated),'rows':len(rows)}
@router.post('/imports/opening-balances')
async def import_balances(user:dict=Depends(admin),file:UploadFile=File(...)):
    rows=rows_from_upload(await file.read(),file.filename or'')
    if not rows:raise HTTPException(400,'The opening-balance file contains no data rows')
    validated=[]
    for index,source in enumerate(rows,2):
        row,values=item_import_values(source,index);missing=[field for field in('warehouse','location','quantity','unit_cost','received_date')if row.get(field)in(None,'')]
        if missing:raise HTTPException(400,f"Row {index}: mandatory opening-balance fields missing: {', '.join(missing)}")
        try:quantity=float(row['quantity']);unit_cost=float(row['unit_cost'])
        except(ValueError,TypeError):raise HTTPException(400,f'Row {index}: quantity and unit_cost must be numeric')
        if quantity<=0 or unit_cost<0:raise HTTPException(400,f'Row {index}: quantity must be positive and unit_cost cannot be negative')
        validated.append((index,row,values,quantity,unit_cost))
    created_items=0;reference=f"OPENING-IMPORT-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    with transaction(immediate=True)as c:
        for index,row,values,quantity,unit_cost in validated:
            warehouse=c.execute('SELECT id FROM warehouses WHERE deleted_at IS NULL AND(lower(name)=lower(?)OR upper(warehouse_code)=upper(?))',(str(row['warehouse']).strip(),str(row['warehouse']).strip())).fetchone()
            if not warehouse:raise HTTPException(400,f'Row {index}: warehouse was not found')
            location=c.execute("SELECT id FROM locations WHERE warehouse_id=? AND deleted_at IS NULL AND type='Bin' AND(lower(code)=lower(?)OR lower(label)=lower(?))",(warehouse['id'],str(row['location']).strip(),str(row['location']).strip())).fetchone()
            if not location:raise HTTPException(400,f'Row {index}: location must identify an active Bin in the selected warehouse')
            item=c.execute('SELECT id FROM items WHERE deleted_at IS NULL AND lower(trim(description))=lower(trim(?))',(values['description'],)).fetchone()
            if item:item_id=item['id']
            else:item_id=insert_import_item(c,values,user['id']);created_items+=1
            receive(c,item_id=item_id,warehouse_id=warehouse['id'],location_id=location['id'],quantity=quantity,unit_cost=unit_cost,batch=str(row.get('batch')or'').strip()or None,expiry_date=str(row.get('expiry_date')or'').strip()or None,received_date=str(row['received_date']).strip(),transaction_type='OPENING_BALANCE',reference_number=reference,reference_table='items',reference_id=item_id,created_by=user['id'])
        log_audit(c,'inventory_stock',0,'CREATE',user['id'],after={'source':'OPENING_BALANCE_IMPORT','reference':reference,'rows':len(validated),'new_items':created_items})
    return {'imported':len(validated),'created':len(validated),'new_items_created':created_items,'rows':len(rows),'reference':reference}
